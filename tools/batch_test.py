"""Batch-test the render endpoint against the social_media_content sheet.

Temporary test harness -- self-contained in this one file so it can be deleted
in a single step when testing is done.

    python -m tools.batch_test              # all rows, 2 at a time
    python -m tools.batch_test --limit 4    # a short trial run first
    python -m tools.batch_test --dry-run    # plan only, no API calls

Reads Text + image from the sheet, renders each row against one platform slot
(rotating through the platforms so the set is covered evenly), writes the
generated images back into the Output Image column, fills the Platform column,
and produces an HTML report so the sheet never has to be opened mid-run.
"""

from __future__ import annotations

import argparse
import asyncio
import base64
import html
import io
import logging
import re
import shutil
import time
import zipfile
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from xml.etree import ElementTree as ET

import httpx
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from PIL import Image

from app.domain.platforms import SPECS, AssetType, Platform, resolve
from app.main import app

SHEET_PATH = Path("social_media_content (2).xlsx")
SHEET_NAME = "Social Media Content"
OUT_DIR = Path("batch_test_output")

COL_TEXT, COL_PLATFORM, COL_INPUT, COL_OUTPUT, COL_URL = 1, 2, 3, 4, 5

# One platform slot per row, rotating so every platform is exercised across the
# set. Website is excluded by request; logo slots take no text overlay.
ROTATION: tuple[tuple[Platform, AssetType], ...] = (
    (Platform.META, AssetType.FEED_SQUARE),
    (Platform.GOOGLE_ADS_PMAX, AssetType.LANDSCAPE),
    (Platform.GOOGLE_BUSINESS_PROFILE, AssetType.PHOTO),
    (Platform.META, AssetType.STORY_REEL),
    (Platform.GOOGLE_ADS_PMAX, AssetType.SQUARE),
    (Platform.META, AssetType.FEED_PORTRAIT),
    (Platform.GOOGLE_ADS_PMAX, AssetType.PORTRAIT),
    (Platform.META, AssetType.FACEBOOK_LANDSCAPE),
)

PLATFORM_LABEL = {
    Platform.META: "Meta",
    Platform.GOOGLE_ADS_PMAX: "Google Ads",
    Platform.GOOGLE_BUSINESS_PROFILE: "GBP",
}

SHAPE_LABEL = {
    AssetType.FEED_SQUARE: "Square",
    AssetType.FEED_PORTRAIT: "Portrait",
    AssetType.STORY_REEL: "Story/Reel",
    AssetType.FACEBOOK_LANDSCAPE: "Landscape",
    AssetType.SQUARE: "Square",
    AssetType.LANDSCAPE: "Landscape",
    AssetType.PORTRAIT: "Portrait",
    AssetType.PHOTO: "Square",
}

NS = {
    "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
    "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
    "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    "rel": "http://schemas.openxmlformats.org/package/2006/relationships",
}

log = logging.getLogger("batch")

# Warnings that describe how the render was produced rather than anything wrong
# with the creative. True and worth keeping in the logs, but they fire on nearly
# every row and bury the findings that matter.
HIDDEN_IN_REPORT = (
    re.compile(r"^Requested \d+x\d+ is not directly renderable"),
    re.compile(r"^The image model returned \d+x\d+; resampled"),
)

# Every static note a platform spec carries ("WebP is not accepted by Google Ads
# PMax", "GBP shows a maximum of 10 photos"). These are properties of the
# platform, identical on every row of that platform, so they say nothing about
# the image in front of you. Pulled from the registry rather than hardcoded, so
# a new platform note never reappears as report noise.
STATIC_PLATFORM_NOTES = {note for spec in SPECS.values() for note in spec.notes}


def visible_warnings(warnings: list[str]) -> list[str]:
    """Warnings that say something about THIS image, not about the platform."""
    return [
        w for w in warnings
        if w not in STATIC_PLATFORM_NOTES
        and not any(p.match(w) for p in HIDDEN_IN_REPORT)
    ]


# --------------------------------------------------------------------------
# reading the sheet


@dataclass
class SourceRow:
    row: int
    text: str
    url: str | None
    embedded: bytes | None
    platform: Platform
    asset_type: AssetType

    @property
    def origin(self) -> str:
        return "url" if self.url else "embedded"

    @property
    def platform_cell(self) -> str:
        spec = resolve(self.platform, self.asset_type)
        width, height = spec.default_size
        return (
            f"{PLATFORM_LABEL[self.platform]} | {width}x{height} | "
            f"{SHAPE_LABEL[self.asset_type]}"
        )


def embedded_images(path: Path) -> dict[int, bytes]:
    """Pictures pasted into the Input Image column, keyed by row number."""
    found: dict[int, bytes] = {}
    with zipfile.ZipFile(path) as archive:
        if "xl/drawings/drawing1.xml" not in archive.namelist():
            return found
        rels = {
            rel.get("Id"): rel.get("Target")
            for rel in ET.fromstring(
                archive.read("xl/drawings/_rels/drawing1.xml.rels")
            ).findall("rel:Relationship", NS)
        }
        for anchor in ET.fromstring(archive.read("xl/drawings/drawing1.xml")):
            origin = anchor.find("xdr:from", NS)
            blip = anchor.find(".//a:blip", NS)
            if origin is None or blip is None:
                continue
            row = int(origin.find("xdr:row", NS).text) + 1
            target = rels.get(blip.get(f"{{{NS['r']}}}embed"), "")
            name = "xl/media/" + target.rsplit("/", 1)[-1]
            if name in archive.namelist():
                found.setdefault(row, archive.read(name))
    return found


def load_rows(path: Path) -> list[SourceRow]:
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[SHEET_NAME]
    pasted = embedded_images(path)

    rows: list[SourceRow] = []
    for number in range(2, sheet.max_row + 1):
        text = sheet.cell(row=number, column=COL_TEXT).value
        if not isinstance(text, str) or not text.strip():
            continue
        raw_url = sheet.cell(row=number, column=COL_URL).value
        url = raw_url if isinstance(raw_url, str) and raw_url.startswith("http") else None
        platform, asset_type = ROTATION[len(rows) % len(ROTATION)]
        rows.append(
            SourceRow(
                row=number,
                text=text.strip(),
                url=url,
                embedded=pasted.get(number),
                platform=platform,
                asset_type=asset_type,
            )
        )
    return rows


# --------------------------------------------------------------------------
# running


# Source photos run to 6000px and several MB. The report needs them legible,
# not original -- 49 full-size files would be ~90MB of page weight.
MAX_VIEW_WIDTH = 900


def display_copy(source: Path, target: Path, max_width: int = MAX_VIEW_WIDTH) -> int:
    """Write a report-sized copy and return the width it should display at.

    Downscales only. A 160px pasted thumbnail is copied untouched and capped at
    160px in the report, so it stays sharp instead of being stretched blurry.
    """
    with Image.open(source) as img:
        img = img.convert("RGB")
        width = min(img.width, max_width)
        if width < img.width:
            img = img.resize(
                (width, round(img.height * width / img.width)), Image.LANCZOS
            )
        target.parent.mkdir(parents=True, exist_ok=True)
        img.save(target, "JPEG", quality=88, optimize=True)
    return width


@dataclass
class Result:
    source: SourceRow
    ok: bool = False
    seconds: float = 0.0
    error: str = ""
    headline: str = ""
    subheadline: str | None = None
    width: int = 0
    height: int = 0
    size_bytes: int = 0
    warnings: list[str] = field(default_factory=list)
    image_path: Path | None = None
    input_path: Path | None = None
    view_path: Path | None = None
    view_width: int = 0


async def source_image(client: httpx.AsyncClient, row: SourceRow) -> bytes:
    if row.embedded:
        return row.embedded
    if not row.url:
        raise RuntimeError("row has neither a pasted image nor a URL")
    response = await client.get(row.url, timeout=60, follow_redirects=True)
    response.raise_for_status()
    return response.content


async def render_row(
    api: httpx.AsyncClient,
    web: httpx.AsyncClient,
    row: SourceRow,
    index: int,
    total: int,
) -> Result:
    result = Result(source=row)
    label = f"[{index:>2}/{total}] row {row.row:<3} {row.platform_cell}"
    started = time.monotonic()
    try:
        image_bytes = await source_image(web, row)
        with Image.open(io.BytesIO(image_bytes)) as img:
            source_size = f"{img.size[0]}x{img.size[1]}"

        result.input_path = OUT_DIR / "input" / f"row{row.row:02d}.jpg"
        result.input_path.parent.mkdir(parents=True, exist_ok=True)
        result.input_path.write_bytes(image_bytes)
        result.view_path = OUT_DIR / "input_view" / f"row{row.row:02d}.jpg"
        result.view_width = display_copy(result.input_path, result.view_path)

        log.info("%s  <- %s %s", label, row.origin, source_size)

        response = await api.post(
            "/api/v1/ad-images/render",
            files={"image": (f"row{row.row}.jpg", image_bytes, "image/jpeg")},
            data={
                "source_text": row.text,
                "platform": row.platform.value,
                "asset_type": row.asset_type.value,
            },
        )
        body = response.json()
        if response.status_code != 200:
            raise RuntimeError(f"{body.get('code', response.status_code)}: "
                               f"{body.get('message', '')[:120]}")

        result.headline = body["ad_copy"]["headline"]
        result.subheadline = body["ad_copy"]["subheadline"]
        result.width = body["image"]["width"]
        result.height = body["image"]["height"]
        result.size_bytes = body["image"]["size_bytes"]
        result.warnings = body["warnings"]

        suffix = body["image"]["image_format"].lower().replace("jpeg", "jpg")
        result.image_path = (
            OUT_DIR / "output"
            / f"row{row.row:02d}_{row.platform.value}_{row.asset_type.value}"
              f"_{result.width}x{result.height}.{suffix}"
        )
        result.image_path.parent.mkdir(parents=True, exist_ok=True)
        result.image_path.write_bytes(base64.b64decode(body["image"]["b64"]))
        result.ok = True
    except Exception as exc:  # noqa: BLE001 - one bad row must not stop the run
        result.error = str(exc)
    result.seconds = time.monotonic() - started

    if result.ok:
        log.info(
            '%s  OK  %dx%d %dKB %.1fs  "%s"%s',
            label, result.width, result.height, result.size_bytes // 1024,
            result.seconds, result.headline,
            f"  ({len(result.warnings)} warning(s))" if result.warnings else "",
        )
    else:
        log.error("%s  FAILED after %.1fs  %s", label, result.seconds, result.error)
    return result


async def run(rows: list[SourceRow], concurrency: int) -> list[Result]:
    limiter = asyncio.Semaphore(concurrency)
    results: list[Result] = []

    async with httpx.AsyncClient(
        transport=httpx.ASGITransport(app=app),
        base_url="http://batch",
        timeout=600,
    ) as api, httpx.AsyncClient() as web:

        async def worker(row: SourceRow, index: int) -> Result:
            async with limiter:
                return await render_row(api, web, row, index, len(rows))

        tasks = [
            asyncio.create_task(worker(row, i))
            for i, row in enumerate(rows, start=1)
        ]
        for task in tasks:
            results.append(await task)
    return results


# --------------------------------------------------------------------------
# writing results back


def thumbnail(path: Path, target: Path, width: int = 150) -> tuple[int, int]:
    with Image.open(path) as img:
        img = img.convert("RGB")
        height = max(1, round(img.height * width / img.width))
        img.resize((width, height), Image.LANCZOS).save(target, "JPEG", quality=82)
    return width, height


def update_workbook(path: Path, results: list[Result]) -> Path:
    backup = path.with_name(
        f"{path.stem}.backup-{datetime.now():%Y%m%d-%H%M%S}{path.suffix}"
    )
    shutil.copy2(path, backup)
    log.info("Backed up the original sheet to %s", backup.name)

    workbook = openpyxl.load_workbook(path)
    sheet = workbook[SHEET_NAME]
    sheet.column_dimensions["B"].width = 34
    sheet.column_dimensions["D"].width = 24

    thumbs = OUT_DIR / "thumbs"
    thumbs.mkdir(parents=True, exist_ok=True)

    written = 0
    for result in results:
        row = result.source.row
        sheet.cell(row=row, column=COL_PLATFORM).value = result.source.platform_cell
        if not (result.ok and result.image_path):
            sheet.cell(row=row, column=COL_OUTPUT).value = f"FAILED: {result.error[:180]}"
            continue

        thumb = thumbs / f"row{row:02d}.jpg"
        width, height = thumbnail(result.image_path, thumb)
        picture = XLImage(str(thumb))
        picture.width, picture.height = width, height
        sheet.add_image(picture, f"D{row}")
        sheet.row_dimensions[row].height = max(
            sheet.row_dimensions[row].height or 0, height * 0.78
        )
        written += 1

    workbook.save(path)
    log.info("Updated %s -- %d image(s) embedded in the Output Image column",
             path.name, written)
    return backup


def write_report(results: list[Result], report: Path) -> None:
    ok = [r for r in results if r.ok]
    failed = [r for r in results if not r.ok]
    total_seconds = sum(r.seconds for r in results)

    def card(result: Result) -> str:
        source = result.source
        if result.ok and result.image_path:
            rel = result.image_path.relative_to(OUT_DIR).as_posix()
            media = f'<img src="{rel}" alt="Generated ad" loading="lazy" />'
            meta = (f"{result.width}×{result.height} · "
                    f"{result.size_bytes // 1024} KB · {result.seconds:.1f}s")
        else:
            media = f'<div class="fail">FAILED<span>{html.escape(result.error)}</span></div>'
            meta = f"{result.seconds:.1f}s"

        input_rel = (result.view_path.relative_to(OUT_DIR).as_posix()
                     if result.view_path else "")
        warnings = "".join(
            f'<li>{html.escape(w)}</li>' for w in visible_warnings(result.warnings)
        )
        return f"""
        <article class="card {'ok' if result.ok else 'bad'}">
          <div class="stage">{media}</div>
          <div class="body">
            <div class="row"><span class="chip">row {source.row}</span>
              <span class="plat">{html.escape(source.platform_cell)}</span></div>
            <p class="headline">{html.escape(result.headline) or "—"}</p>
            {f'<p class="sub">{html.escape(result.subheadline)}</p>' if result.subheadline else ''}
            <p class="meta">{meta}</p>
            {f'<ul class="warn">{warnings}</ul>' if warnings else ''}
            <details><summary>Source text &amp; input</summary>
              <p class="src">{html.escape(source.text)}</p>
              {f'<img class="thumbin" style="max-width:{result.view_width}px" src="{input_rel}" alt="Source image" loading="lazy" />' if input_rel else ''}
            </details>
          </div>
        </article>"""

    report.write_text(f"""<!doctype html>
<html lang="en"><head><meta charset="utf-8" />
<meta name="viewport" content="width=device-width, initial-scale=1" />
<title>Batch Test Results</title>
<style>
  :root {{
    --paper:#f7f8fa; --surface:#fff; --ink:#16202e; --soft:#5a6878;
    --faint:#8b97a6; --rule:#dfe4eb; --accent:#1d5bd6; --ok:#067647;
    --warn:#a8620a; --warn-dim:#fdf3e6; --bad:#b42318; --bad-dim:#fdecea;
    --mono:ui-monospace,"Cascadia Mono",Consolas,monospace;
  }}
  *{{box-sizing:border-box}}
  body{{margin:0;background:var(--paper);color:var(--ink);
    font:15px/1.55 ui-sans-serif,"Segoe UI",system-ui,sans-serif}}
  header{{padding:22px 28px;background:var(--surface);
    border-bottom:1px solid var(--rule);display:flex;flex-wrap:wrap;
    gap:20px;align-items:baseline;justify-content:space-between}}
  h1{{margin:0;font-size:17px;font-weight:650}}
  .stats{{display:flex;gap:18px;font-family:var(--mono);font-size:12.5px}}
  .stats b{{font-size:15px}}
  .pass{{color:var(--ok)}} .failc{{color:var(--bad)}}
  .grid{{display:grid;gap:16px;padding:24px 28px;
    grid-template-columns:repeat(auto-fill,minmax(300px,1fr))}}
  .card{{background:var(--surface);border:1px solid var(--rule);border-radius:8px;
    overflow:hidden;display:flex;flex-direction:column;
    box-shadow:0 1px 2px rgba(22,32,46,.05)}}
  .card.bad{{border-color:#f0c4c0}}
  .stage{{background:#f2f4f8;display:grid;place-items:center;padding:14px;min-height:190px}}
  .stage img{{max-width:100%;max-height:300px;border:1px solid var(--rule)}}
  .fail{{display:flex;flex-direction:column;gap:6px;align-items:center;
    color:var(--bad);font:600 13px var(--mono);text-align:center}}
  .fail span{{font-weight:400;font-size:11.5px;max-width:34ch}}
  .body{{padding:12px 14px;display:flex;flex-direction:column;gap:7px}}
  .row{{display:flex;gap:8px;align-items:center;flex-wrap:wrap}}
  .chip{{font:600 10px var(--mono);letter-spacing:.06em;text-transform:uppercase;
    background:#eef2f9;color:var(--soft);padding:3px 7px;border-radius:100px}}
  .plat{{font:500 11.5px var(--mono);color:var(--accent)}}
  .headline{{margin:0;font-size:15px;font-weight:640;line-height:1.3}}
  .sub{{margin:0;font-size:13px;color:var(--soft)}}
  .meta{{margin:0;font:400 11.5px var(--mono);color:var(--faint)}}
  .warn{{margin:2px 0 0;padding:8px 10px 8px 22px;background:var(--warn-dim);
    border-left:3px solid var(--warn);border-radius:4px;
    font-size:11.5px;color:#6f4207}}
  details{{font-size:12px;color:var(--soft)}}
  summary{{cursor:pointer;color:var(--accent);font-size:12px}}
  .src{{margin:6px 0;line-height:1.5}}
  .thumbin{{display:block;width:100%;height:auto;margin-top:8px;
    border:1px solid var(--rule);border-radius:4px;background:#f2f4f8}}
</style></head><body>
<header>
  <div><h1>Batch Test Results</h1>
    <p style="margin:3px 0 0;font-size:13px;color:var(--soft)">
      {datetime.now():%d %b %Y, %H:%M} · sheet: {html.escape(SHEET_PATH.name)}</p></div>
  <div class="stats">
    <span>total <b>{len(results)}</b></span>
    <span class="pass">passed <b>{len(ok)}</b></span>
    <span class="failc">failed <b>{len(failed)}</b></span>
    <span>time <b>{total_seconds / 60:.1f}m</b></span>
  </div>
</header>
<div class="grid">{''.join(card(r) for r in results)}</div>
</body></html>""", encoding="utf-8")
    log.info("Report written to %s", report)


# --------------------------------------------------------------------------


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--limit", type=int, help="only process the first N rows")
    parser.add_argument("--concurrency", type=int, default=2,
                        help="rows processed at once (default 2)")
    parser.add_argument("--dry-run", action="store_true",
                        help="show the plan without calling any API")
    parser.add_argument("--yes", action="store_true", help="skip the confirmation")
    args = parser.parse_args()

    logging.basicConfig(
        level=logging.INFO, format="%(asctime)s  %(message)s", datefmt="%H:%M:%S"
    )
    logging.getLogger("httpx").setLevel(logging.WARNING)

    if not SHEET_PATH.exists():
        log.error("Sheet not found: %s", SHEET_PATH)
        return 1

    rows = load_rows(SHEET_PATH)
    if args.limit:
        rows = rows[: args.limit]

    log.info("Loaded %d row(s) from %s", len(rows), SHEET_PATH.name)
    counts: dict[str, int] = {}
    for row in rows:
        counts[row.platform_cell] = counts.get(row.platform_cell, 0) + 1
    for cell, count in sorted(counts.items()):
        log.info("   %-36s %d row(s)", cell, count)
    log.info("   %d from URL, %d from pasted image",
             sum(1 for r in rows if r.url), sum(1 for r in rows if not r.url))

    if args.dry_run:
        log.info("Dry run -- nothing was called and nothing was written.")
        return 0

    if not args.yes:
        print(f"\nThis will make {len(rows)} copy calls and {len(rows)} image "
              f"renders against the live OpenAI API, and overwrite "
              f"'{SHEET_PATH.name}' (a timestamped backup is written first).")
        if input("Continue? [y/N] ").strip().lower() not in {"y", "yes"}:
            log.info("Aborted.")
            return 1

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    started = time.monotonic()
    results = asyncio.run(run(rows, args.concurrency))
    elapsed = time.monotonic() - started

    passed = sum(1 for r in results if r.ok)
    log.info("-" * 68)
    log.info("Done in %.1f min -- %d passed, %d failed",
             elapsed / 60, passed, len(results) - passed)

    update_workbook(SHEET_PATH, results)
    report = OUT_DIR / "report.html"
    write_report(results, report)
    log.info("Open the report: %s", report.resolve())
    return 0 if passed == len(results) else 2


if __name__ == "__main__":
    raise SystemExit(main())
